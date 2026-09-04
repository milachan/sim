# Analisis konvensi jam: jam slot = kolom MAPEL atau kolom KELAS?
# Uji 1: konflik kelas (kelas sama di hari+jam sama dari guru berbeda) harus 0 utk hipotesis benar.
# Uji 2: cocokkan dengan xlsx Selatan utk guru yang sama (hari+jam+mapel).
import json
import re
from collections import Counter, defaultdict
from pypdf import PdfReader

PDF = "contoh_jadwal/RALAT PER GURU UTARA_SEPTEMBER.pdf"
XLSX = "contoh_jadwal/Jadwal_Guru_Selatan_Agustus_Terstruktur.xlsx"

NAMA_HARI = ["Senin", "Selasa", "Rabu", "Kamis", "Jumat", "Sabtu"]
LABEL_HARI = {"Se", "Ra", "Ka", "Ju", "Sa"}
TINGKAT = {"VII", "VIII", "IX"}
RE_KELAS = re.compile(r"^(VII|VIII|IX)\s+([A-Z])$")


def extract_words(reader, page_idx):
    page = reader.pages[page_idx]
    raw = []

    def visitor(text, cm, tm, font_dict, font_size):
        if not text or not text.strip():
            return
        x0 = tm[4]
        y0 = tm[5]
        sx = tm[0] if tm[0] else 1.0
        raw.append((x0, y0, text, sx))

    page.extract_text(visitor_text=visitor)
    raw.sort(key=lambda c: (-c[1], c[0]))
    words = []
    cur = None
    for (x, y, text, sx) in raw:
        w = len(text) * max(sx, 0.1)
        if cur is None or abs(y - cur["y"]) > 3.0 or x > cur["x1"] + 2.0:
            if cur:
                words.append(cur)
            cur = {"x0": x, "x1": x + w, "y": y, "t": text}
        else:
            cur["x1"] = max(cur["x1"], x + w)
            cur["t"] += text
    if cur:
        words.append(cur)
    return words


def kolom_dari_x(x_mid, centers):
    """Kolom 1..9 dari posisi x (tengah) memakai batas antarkolom (titik tengah antar center)."""
    if not centers:
        return None
    # batas = titik tengah antara center berurutan; kolom kiri/batas luar diperluas
    bounds = [centers[0] - (centers[1] - centers[0]) / 2]
    for i in range(len(centers) - 1):
        bounds.append((centers[i] + centers[i + 1]) / 2)
    bounds.append(centers[-1] + (centers[-1] - centers[-2]) / 2)
    for i in range(len(bounds) - 1):
        if bounds[i] <= x_mid < bounds[i + 1]:
            return i + 1
    return None


def frasa_mapel(tokens):
    """Gabung token berjarak dekat menjadi frasa (satu mapel)."""
    out = []
    for w in tokens:
        if out and w["x0"] - out[-1]["x1"] <= 10:
            out[-1]["t"] += " " + w["t"]
            out[-1]["x1"] = max(out[-1]["x1"], w["x1"])
        else:
            out.append({"t": w["t"], "x0": w["x0"], "x1": w["x1"]})
    return out


def kelas_dari_tokens(tokens):
    """Daftar kelas {t, x0, x1} dari token tingkat+huruf (gabung atau terpisah)."""
    kelas = []
    used = set()
    # 1) token gabung "IX K"
    for i, w in enumerate(tokens):
        if i in used:
            continue
        m = RE_KELAS.match(w["t"])
        if m:
            kelas.append({"t": f"{m.group(1)} {m.group(2)}", "x0": w["x0"], "x1": w["x1"]})
            used.add(i)
    # 2) token terpisah "IX" ... "K" (dalam radius 30)
    for i, w in enumerate(tokens):
        if i in used or w["t"] not in TINGKAT:
            continue
        for j in range(i + 1, len(tokens)):
            if j in used:
                continue
            if re.fullmatch(r"[A-Z]", tokens[j]["t"]) and tokens[j]["x0"] - w["x1"] <= 30:
                kelas.append({"t": f"{w['t']} {tokens[j]['t']}", "x0": w["x0"], "x1": tokens[j]["x1"]})
                used.add(i)
                used.add(j)
                break
            if tokens[j]["x0"] - w["x1"] > 30:
                break
    return kelas


def parse_semua():
    reader = PdfReader(PDF)
    hasil = []  # {guru, kode, hari, mapel, mcol, kelas, kcol}
    warnings = []
    for pi in range(len(reader.pages)):
        words = extract_words(reader, pi)
        if not words:
            warnings.append(f"hal {pi+1}: kosong")
            continue
        # baris per y asc (atas dulu)
        baris = []
        for w in sorted(words, key=lambda w: (w["y"], w["x0"])):
            if baris and abs(w["y"] - baris[-1]["y"]) <= 5.0:
                baris[-1]["w"].append(w)
            else:
                baris.append({"y": w["y"], "w": [w]})
        for b in baris:
            b["w"].sort(key=lambda w: w["x0"])

        guru = kode = None
        for b in baris:
            teks = " ".join(w["t"] for w in b["w"])
            m = re.search(r"Teacher\s+(.+?)\s*\(([A-Z]+\d+)\)", teks)
            if m:
                guru = re.sub(r"\s+", " ", m.group(1)).strip()
                kode = m.group(2)
                break
        if not guru:
            warnings.append(f"hal {pi+1}: guru tak ditemukan")
            continue

        waktu = None
        for b in baris:
            teks = " ".join(w["t"] for w in b["w"])
            if "6:50" in teks and "7:55" in teks:
                waktu = b
                break
        if not waktu:
            warnings.append(f"hal {pi+1} ({kode}): waktu tak ditemukan")
            continue
        centers = []
        for w in waktu["w"]:
            for m in re.finditer(r"\d{1,2}:\d{2}\s*-\s*\d{1,2}:\d{2}", w["t"]):
                sp = m.end() - m.start()
                fr = (m.start() + sp / 2) / len(w["t"])
                centers.append(w["x0"] + fr * (w["x1"] - w["x0"]))
        if len(centers) < 9:
            # coba interpolar: beberapa kolom mungkin dalam satu kata berjarak sama? abaikan dulu
            warnings.append(f"hal {pi+1} ({kode}): hanya {len(centers)} kolom waktu")

        hariCounter = -1
        pendingMapel = None
        for b in baris:
            if b is waktu:
                continue
            teks = " ".join(w["t"] for w in b["w"])
            if teks.startswith("Menghasilkan") or "aSc Timetables" in teks:
                continue
            if "Teacher" in teks or teks.strip() == "IKA MATSDA":
                continue
            if re.fullmatch(r"[\d\s]+", teks):
                continue  # baris nomor kolom 1..9

            first = b["w"][0] if b["w"] else None
            isLabel = first is not None and first["t"] in LABEL_HARI and first["x0"] < 120
            if isLabel:
                hariCounter += 1
            sisa = b["w"][1:] if isLabel else b["w"]

            kls = kelas_dari_tokens(sisa)
            if kls:
                hari = NAMA_HARI[hariCounter] if 0 <= hariCounter < 6 else f"?{hariCounter+1}"
                if pendingMapel:
                    if len(pendingMapel) != len(kls):
                        warnings.append(f"hal {pi+1} ({kode}) {hari}: {len(pendingMapel)} mapel vs {len(kls)} kelas")
                    for i in range(min(len(pendingMapel), len(kls))):
                        mcol = kolom_dari_x((pendingMapel[i]["x0"] + pendingMapel[i]["x1"]) / 2, centers)
                        kcol = kolom_dari_x((kls[i]["x0"] + kls[i]["x1"]) / 2, centers)
                        hasil.append({
                            "guru": guru, "kode": kode, "hari": hari,
                            "mapel": pendingMapel[i]["t"], "mcol": mcol,
                            "kelas": kls[i]["t"], "kcol": kcol,
                        })
                else:
                    warnings.append(f"hal {pi+1} ({kode}) {hari}: kelas tanpa mapel sebelumnya")
                pendingMapel = None
            else:
                mp = [w for w in sisa if w["t"] not in LABEL_HARI]
                if mp:
                    pendingMapel = frasa_mapel(mp)
    return hasil, warnings


def main():
    hasil, warnings = parse_semua()
    print(f"TOTAL SLOT: {len(hasil)}")
    print(f"WARNING: {len(warnings)}")
    for w in warnings[:40]:
        print("  WARN:", w)

    # Uji 1: konflik KELAS (kelas sama di hari+jam sama = 2 guru klaim) & GURU (guru sama di hari+jam sama = mustahil)
    for nama, key in [("H0=kolom-mapel", "mcol"), ("H1=kolom-kelas", "kcol")]:
        ck = Counter((r["kelas"], r["hari"], r[key]) for r in hasil if r[key])
        cg = Counter((r["guru"], r["hari"], r[key]) for r in hasil if r[key])
        konflikKelas = {k: v for k, v in ck.items() if v > 1}
        konflikGuru = {k: v for k, v in cg.items() if v > 1}
        print(f"\n[{nama}] konflik kelas x2+: {len(konflikKelas)} | konflik GURU x2+: {len(konflikGuru)}")
        for k, v in sorted(konflikGuru.items()):
            print(f"   GURU {k} x{v} -> {[(r['kelas'], r['mapel']) for r in hasil if r['guru']==k[0] and r['hari']==k[1] and r[key]==k[2]]}")

    # Uji 2: bandingkan dengan xlsx Selatan (16 guru sama)
    try:
        from openpyxl import load_workbook
        wb = load_workbook(XLSX, read_only=True, data_only=True)
        ws = wb.active
        sel = defaultdict(list)  # kode -> [(hari, jam, mapel, kelas)]
        for row in ws.iter_rows(values_only=True):
            g, hari, jam, waktu, mapel, kelas = row[:6]
            if not g or not hari:
                continue
            m = re.search(r"\(([A-Z]+\d+)\)", str(g))
            if not m:
                continue
            sel[m.group(1)].append((str(hari).strip(), int(jam), str(mapel).strip(), str(kelas).strip()))
        shared = set(sel) & {r["kode"] for r in hasil}
        print(f"\nGuru bersama utk uji xlsx: {len(shared)}")

        def norm(s):
            return re.sub(r"[^a-z0-9]", "", s.lower())

        for nama, key in [("H0=kolom-mapel", "mcol"), ("H1=kolom-kelas", "kcol")]:
            cocok = 0
            total = 0
            for kode in sorted(shared):
                utara = [(r["hari"], r[key], norm(r["mapel"])) for r in hasil if r["kode"] == kode and r[key]]
                # utk tiap baris selatan, cari padanan utara dgn (hari, mapel) sama; cocok = jam sama
                for (hari, jam, mapel, kelas) in sel[kode]:
                    mn = norm(mapel)
                    cands = [(h, j) for (h, j, m) in utara if h == hari and m == mn]
                    if not cands:
                        continue
                    total += len(cands)
                    cocok += sum(1 for (_, j) in cands if j == jam)
            print(f"[{nama}] cocok jam (hari+mapel sama): {cocok}/{total}")

        # tampilkan per guru utk inspeksi
        print("\n--- PERBANDINGAN PER GURU (H0 vs H1 vs xlsx) ---")
        for kode in sorted(shared):
            print(f"\n== {kode} ==")
            print("  xlsx Selatan:", sel[kode])
            u0 = sorted((r["hari"], r["mcol"], r["mapel"], r["kelas"]) for r in hasil if r["kode"] == kode)
            u1 = sorted((r["hari"], r["kcol"], r["mapel"], r["kelas"]) for r in hasil if r["kode"] == kode)
            print("  utara H0 (jam=mapel):", u0)
            print("  utara H1 (jam=kelas):", u1)
    except ImportError:
        print("\n(openpyxl tidak ada; lewati uji xlsx)")

    with open("contoh_jadwal/slot-utara.json", "w", encoding="utf-8") as f:
        json.dump(hasil, f, ensure_ascii=False, indent=1)
    print("\nOK → contoh_jadwal/slot-utara.json")


if __name__ == "__main__":
    main()